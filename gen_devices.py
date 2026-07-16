"""Generate src/devices.js from Config_actelis-price-list.xlsx + AutoRepeaterInfo.xlsx.

Why this exists
---------------
The device list used to be hand-written, which meant a product could be on the
price list, have full compatibility rules in AutoRepeaterInfo, and still not
appear in the tool — nobody had typed it in. ML6916EL was exactly that. Deriving
the list from the price list makes that class of bug impossible: if it's a
device category on the price list, it's in the tool.

Run:  python3 gen_devices.py
"""
import openpyxl, re, json

PRICES = '/mnt/user-data/uploads/Config_actelis-price-list.xlsx'
ARI    = '/mnt/user-data/uploads/AutoRepeaterInfo.xlsx'
OUT    = '/home/claude/actelis-accessory-selector/src/devices.js'

# Only these price-list categories are selectable products. Accessories (SFPs,
# cables, PSUs) are offered by the compatibility engine, not chosen up-front.
CAT_META = {
    "A1.1 GL800 Solutions":              ("GL800",     "standalone", "node"),
    "A1.2 GL9000 Headend Solutions":     ("GL9000HE",  "rack",       "rack"),
    "A1.3 GL9000 CPEs":                  ("GL9000CPE", "cpe",        "cpe"),
    "A1.4 GL900 Headend Solutions":      ("GL900HE",   "standalone", "node"),
    "A1.5 GL900 CPEs":                   ("GL900CPE",  "cpe",        "cpe"),
    "A2. ML600 Family":                  ("ML600",     "standalone", "node"),
    "A2.1 ML600D Family":                ("ML600D",    "din",        "din"),
    "A4.1 Shelves":                      ("CHASSIS",   "chassis",    "chassis"),
    "A4.2 SDU (Service Dispatch Units)": ("CHASSIS",   "chassis",    "card"),
    "A4.3 MLU (Multiport Line Units)":   ("CHASSIS",   "chassis",    "card"),
    "A5.1 Fiber DIN Rail L2 Switches":   ("GL5000",    "din",        "din"),
    "A5.2 Fiber DIN Rail L3 Switches":   ("GL5000",    "din",        "din"),
    "A5.3 Fiber Rackmount L2 Switches":  ("GL6000",    "rack",       "rack"),
    "A5.4 Fiber Rackmount L3 Switches":  ("GL6000",    "rack",       "rack"),
    "A5.5 Fiber In-Pole L2 Switches":    ("GL7000",    "inpole",     "pole"),
    "A9. L3 CPEs":                       ("CPE",       "cpe",        "cpe"),
}

# ── AutoRepeaterInfo (NA rows) ───────────────────────────────────────────────
wb = openpyxl.load_workbook(ARI, data_only=True)
ws = wb['AutoRepeaterInfo']
rows = list(ws.iter_rows(values_only=True))
hdr = [str(h).strip() for h in rows[0]]
ari = {}
for r in rows[1:]:
    if not any(r) or not r[0]:
        continue
    d = dict(zip(hdr, r))
    reg = d['Region']
    is_na = reg is None or str(reg) == 'None' or 'NA' in str(reg).split()
    if not is_na:
        continue
    ari.setdefault(str(d['Part Number']).strip(), d)

# ── Price list ───────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(PRICES, data_only=True)
ws = wb['Sheet1']
devices, seen = [], set()

def clean_model(desc, cat):
    """Turn the price list's Description into a usable model name.

    The column is inconsistent: sometimes it IS the model ("GL830-8O"),
    sometimes prose with the model in brackets ("Service Dispatcher Unit
    (SDU-450)"), sometimes prose with no model at all ("Chassis 200 Shelf").
    For shelves we use the Access wizard's own identification, which matches on
    exactly these description strings (Form_Wizard-Node.cls, CO_Model_AfterUpdate).
    """
    d = (desc or '').strip()

    if cat.startswith('A4.1'):
        if 'Chassis 100 Shelf' in d:
            return 'CHS-100'
        if 'Chassis 200 Shelf' in d:
            return 'CHS-200'
        if 'Chassis 2000' in d:
            return 'CHS-2000B' if '2000B' in d else 'CHS-2000'

    m = re.match(r'^(.*?)\s*\((.*)\)\s*$', d)
    if m:
        inner = m.group(2).strip()
        # A bracketed part code is the real model: "(SDU-450)", "(CHS-2000B/19)".
        # Require a digit so "(New)" and "(w/o SyncE)" — which distinguish two
        # otherwise identical models — are kept as part of the name instead.
        if len(inner) <= 16 and re.fullmatch(r'[A-Z][A-Za-z0-9./-]*\d[A-Za-z0-9./-]*', inner):
            return inner
        if len(inner) > 24:
            return m.group(1).strip()
    return d

def short(comment, limit=76):
    """A card-sized summary. The Comments column runs to 250+ characters; cut on
    a comma so the fragment still reads as a phrase."""
    c = re.sub(r'\s+', ' ', (comment or '')).strip()
    if len(c) <= limit:
        return c
    cut = c[:limit]
    i = cut.rfind(',')
    return (cut[:i] if i > 30 else cut).rstrip(' ,') + '…'

def iface_for(cat, comment, model):
    """Interface tags. Only consulted for devices absent from AutoRepeaterInfo —
    everything in ARI is driven by its real columns instead. Derived from the
    price-list text, so treat as a hint, not a spec."""
    c = (comment or '') + ' ' + model
    lo = c.lower()
    gl9000 = cat.startswith(('A1.2', 'A1.3'))
    gl900  = cat.startswith(('A1.4', 'A1.5'))
    coax   = gl9000 or bool(re.search(r'g\.hn|coax', lo)) and not gl900
    copper = bool(re.search(r'\d+\s*pairs?', lo)) or gl900 or cat.startswith('A9')
    fiber  = bool(re.search(r'sfp|base-fx|base-f\b', lo))
    poe    = bool(re.search(r'\bpoe\b|-p\b', lo))
    return {"copper": bool(copper and not coax), "coax": bool(coax),
            "fiber": bool(fiber), "poe": bool(poe)}

for r in ws.iter_rows(min_row=2, values_only=True):
    pn = r[1]
    if not pn:
        continue
    pn = str(pn).strip()
    cat = str(r[0]).strip() if r[0] else ''
    if cat not in CAT_META or pn in seen:
        continue
    seen.add(pn)
    family, form, icon = CAT_META[cat]
    model = clean_model(str(r[2]).strip() if r[2] else pn, cat)
    spec = re.sub(r'\s+', ' ', str(r[3]).strip()) if r[3] else ''

    devices.append({
        "pn": pn,
        "model": model,
        "desc": short(spec) or model,
        "spec": spec,
        "family": family,
        "form": form,
        "icon": icon,
        "iface": iface_for(cat, spec, model),
        "ari": pn in ari,
    })

devices.sort(key=lambda d: (d['family'], d['model']))

body = ',\n'.join(
    '  { pn:"%s", model:%s, family:"%s", form:"%s", icon:"%s",\n'
    '    desc:%s,\n'
    '    spec:%s,\n'
    '    iface:%s }'
    % (d['pn'], json.dumps(d['model']), d['family'], d['form'], d['icon'],
       json.dumps(d['desc']), json.dumps(d['spec']),
       json.dumps(d['iface']).replace('"copper"', 'copper').replace('"coax"', 'coax')
                             .replace('"fiber"', 'fiber').replace('"poe"', 'poe')
                             .replace('true', 'true').replace(': ', ':'))
    for d in devices
)

hdr_txt = '''// ════════════════════════════════════════════════════════════════════════════
//  DEVICES — GENERATED from the price list. Do not hand-edit.
//
//  Every part in a device category of Config_actelis-price-list.xlsx appears
//  here, so a product can't be missing from the tool just because nobody typed
//  it in. Regenerate with gen_devices.py when a new price list lands.
//
//    model / desc / spec   from the price list (desc is the card-sized summary,
//                          spec the full text shown in the detail view)
//    family / form / icon  from the part's price-list category
//    iface                 a hint derived from the price-list text, used ONLY
//                          for devices absent from AutoRepeaterInfo; everything
//                          in ARI is driven by its real columns instead
//
//  Price, datasheet and filter tab are attached in catalog.js from the price
//  list — they are not repeated here.
// ════════════════════════════════════════════════════════════════════════════

export const DEVICE_ROWS = [
'''

with open(OUT, 'w') as f:
    f.write(hdr_txt + body + ',\n];\n')

n_ari = sum(1 for d in devices if d['ari'])
print('devices: %d  (%d with AutoRepeaterInfo rules, %d inferred)'
      % (len(devices), n_ari, len(devices) - n_ari))
from collections import Counter
for k, v in sorted(Counter(d['family'] for d in devices).items()):
    print('   %-10s %d' % (k, v))
